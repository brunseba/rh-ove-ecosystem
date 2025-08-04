#!/usr/bin/env python3
"""
Export weekly workload breakdown to multi-sheet XLSX file
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

def parse_markdown_tables(file_path):
    """Parse markdown file and extract tables"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Define data structures for each sheet
    sheets_data = {}
    
    # Persona Skills and Availability
    persona_data = [
        ['Persona Type', 'Skill Level', 'Standard Days/Week', 'Notes'],
        ['Infrastructure Architect', 'Senior', 5, 'Part-time (2.5 days) during Implementation'],
        ['Solution Architect', 'Senior', 5, 'Full-time during active phases'],
        ['Migration Specialist', 'Senior', 5, 'Full-time throughout migration project'],
        ['DevOps Engineer', 'Mid-Senior', 5, 'Multiple resources (2x)'],
        ['Security Engineer', 'Senior', 5, 'Full-time during active phases'],
        ['Network Engineer', 'Senior', 5, 'Full-time during active phases'],
        ['Security Specialist', 'Senior', 5, 'Full-time during active phases'],
        ['Application Owner', 'Mid-Senior', 5, 'Multiple resources (2x)'],
        ['Performance Engineer', 'Mid-Senior', 5, 'Full-time during active phases'],
        ['Business Analyst', 'Mid-Senior', 5, 'Full-time during active phases'],
        ['RH OVE Engineer', 'Mid-Senior', 5, 'Multiple resources (2x)'],
        ['Application Developer', 'Mid-Level', 5, 'Multiple resources (3x)'],
        ['System Administrator', 'Mid-Level', 5, 'Multiple resources (2x)'],
        ['Testing Specialist', 'Mid-Level', 5, 'Full-time during active phases'],
        ['VMware Administrator', 'Mid-Level', 5, 'Multiple resources (2x)'],
        ['Backup Administrator', 'Mid-Level', 5, 'Full-time during active phases']
    ]
    sheets_data['Personas'] = persona_data
    
    # Infrastructure Project - Weekly Allocation
    infra_weekly = [
        ['Week', 'Phase', 'Infrastructure Architect', 'DevOps Engineer (2x)', 'System Administrator (2x)', 'Security Engineer', 'Network Engineer', 'Weekly Total (Days)'],
        ['1-4', 'Study', 5, 0, 0, 0, 5, 10],
        ['5-12', 'Design', 5, 10, 0, 5, 5, 25],
        ['13-20', 'Implementation', 2.5, 10, 10, 5, 5, 32.5],
        ['21-24', 'Testing', 0, 10, 10, 5, 0, 25],
        ['25-28', 'Day-2 Ops', 0, 10, 10, 0, 0, 20]
    ]
    sheets_data['Infrastructure Weekly'] = infra_weekly
    
    # Infrastructure Project - Phase Summary
    infra_phases = [
        ['Phase', 'Weeks', 'Person-Days'],
        ['Study Phase', '1-4', 40],
        ['Design Phase', '5-12', 200],
        ['Implementation Phase', '13-20', 260],
        ['Testing Phase', '21-24', 100],
        ['Day-2 Operations', '25-28', 80],
        ['TOTAL', '', 680]
    ]
    sheets_data['Infrastructure Phases'] = infra_phases
    
    # Use-Cases Project - Weekly Allocation
    usecase_weekly = [
        ['Week', 'Phase', 'Solution Architect', 'Application Developer (3x)', 'Testing Specialist', 'DevOps Engineer', 'Security Specialist', 'Business Analyst', 'Weekly Total (Days)'],
        ['15-18', 'Study', 5, 0, 0, 0, 0, 5, 10],
        ['19-26', 'Design', 5, 0, 0, 0, 5, 5, 15],
        ['27-38', 'Implementation', 0, 15, 0, 5, 5, 0, 25],
        ['39-42', 'Testing', 0, 15, 5, 5, 5, 0, 30],
        ['43-46', 'Day-2 Ops', 0, 0, 0, 5, 0, 0, 5]
    ]
    sheets_data['Use-Cases Weekly'] = usecase_weekly
    
    # Use-Cases Project - Phase Summary
    usecase_phases = [
        ['Phase', 'Weeks', 'Person-Days'],
        ['Study Phase', '15-18', 40],
        ['Design Phase', '19-26', 120],
        ['Implementation Phase', '27-38', 300],
        ['Testing Phase', '39-42', 120],
        ['Day-2 Operations', '43-46', 20],
        ['TOTAL', '', 600]
    ]
    sheets_data['Use-Cases Phases'] = usecase_phases
    
    # Migration Project - Weekly Allocation
    migration_weekly = [
        ['Week', 'Phase', 'Migration Specialist', 'VMware Admin (2x)', 'RH OVE Engineer (2x)', 'Application Owner (2x)', 'Performance Engineer', 'Backup Administrator', 'Weekly Total (Days)'],
        ['21-28', 'Study', 5, 10, 0, 0, 0, 0, 15],
        ['29-36', 'Design', 5, 10, 0, 10, 0, 0, 25],
        ['37-50', 'Implementation', 5, 10, 10, 10, 5, 5, 45],
        ['51-54', 'Testing', 5, 0, 10, 10, 5, 5, 35],
        ['55-60', 'Day-2 Ops', 5, 0, 10, 0, 5, 5, 25]
    ]
    sheets_data['Migration Weekly'] = migration_weekly
    
    # Migration Project - Phase Summary
    migration_phases = [
        ['Phase', 'Weeks', 'Person-Days'],
        ['Study Phase', '21-28', 120],
        ['Design Phase', '29-36', 200],
        ['Implementation Phase', '37-50', 630],
        ['Testing Phase', '51-54', 140],
        ['Day-2 Operations', '55-60', 150],
        ['TOTAL', '', 1240]
    ]
    sheets_data['Migration Phases'] = migration_phases
    
    # Project Summary
    project_summary = [
        ['Sub-Project', 'Duration (Weeks)', 'Total Workload (Person-Days)', 'Average Weekly Workload (Days)'],
        ['RH OVE Infrastructure', 28, 582.5, 20.8],
        ['Use-Cases Implementation', 32, 560, 17.5],
        ['Migration from VMware', 40, 869, 21.7],
        ['TOTAL', 100, 2011.5, 60.0]
    ]
    sheets_data['Project Summary'] = project_summary
    
    # Peak Resource Utilization
    peak_utilization = [
        ['Week Range', 'Projects Active', 'Weekly Workload (Days)', 'Key Activities'],
        ['37-42', 'Infrastructure + Use-Cases + Migration', 102.5, 'Implementation overlap'],
        ['43-46', 'Use-Cases + Migration', 30, 'Testing phases'],
        ['47-50', 'Migration Implementation', 45, 'Critical migration waves']
    ]
    sheets_data['Peak Utilization'] = peak_utilization
    
    # Workload Distribution by Persona
    workload_distribution = [
        ['Persona Type', 'Total Weeks Active', 'Total Workload (Days)', 'Percentage'],
        ['Migration Specialist', 40, 200, 9.9],
        ['Application Developer (3x)', 16, 240, 11.9],
        ['RH OVE Engineer (2x)', 24, 240, 11.9],
        ['VMware Administrator (2x)', 28, 280, 13.9],
        ['DevOps Engineer (2x)', 36, 360, 17.9],
        ['Infrastructure Architect', 18.5, 92.5, 4.6],
        ['System Administrator (2x)', 16, 160, 8.0],
        ['Security Engineer', 16, 80, 4.0],
        ['Network Engineer', 20, 100, 5.0],
        ['Solution Architect', 12, 60, 3.0],
        ['Testing Specialist', 4, 20, 1.0],
        ['Security Specialist', 12, 60, 3.0],
        ['Business Analyst', 4, 20, 1.0],
        ['Application Owner (2x)', 26, 260, 12.9],
        ['Performance Engineer', 18, 90, 4.5],
        ['Backup Administrator', 30, 150, 7.5]
    ]
    sheets_data['Workload Distribution'] = workload_distribution
    
    return sheets_data

def create_xlsx_workbook(data, output_file):
    """Create XLSX workbook with multiple sheets"""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for sheet_name, sheet_data in data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Add data to worksheet
        for row_idx, row_data in enumerate(sheet_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Style header row
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Style total rows
                if isinstance(cell_value, str) and ('TOTAL' in cell_value.upper() or 'Total' in str(cell_value)):
                    cell.font = total_font
                    cell.fill = total_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"XLSX file created: {output_file}")

def main():
    markdown_file = "../docs/project-plan/weekly-charge-breakdown.md"
    xlsx_file = "../docs/export/RH_OVE_Weekly_Workload_Breakdown.xlsx"
    
    print("Parsing workload data from markdown file...")
    data = parse_markdown_tables(markdown_file)
    
    print("Creating XLSX workbook with multiple sheets...")
    create_xlsx_workbook(data, xlsx_file)
    
    print(f"\nExport complete! Multi-sheet XLSX file created: {xlsx_file}")
    print(f"Sheets included:")
    for sheet_name in data.keys():
        print(f"  - {sheet_name}")

if __name__ == "__main__":
    main()
